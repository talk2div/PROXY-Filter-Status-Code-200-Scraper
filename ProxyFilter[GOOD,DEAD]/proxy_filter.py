import requests
import openpyxl
ip_addresses = []
good = []
dead = []
i = 0
workbook = openpyxl.load_workbook("E://Web Scraping Project//freeproxy//ProxyFilter[GOOD,DEAD]//proxy_filter_file.xlsx")
sheet = workbook["PROXY"]

# fetching proxy column and saving into ip list
row_count = sheet.max_row
for row in range(2,row_count+1):
    ip_addresses.append(sheet.cell(row=row,column=1).value)

# checking all proxy having status code = 200
retry = 1
while i < len(ip_addresses):
    try:
        proxy = {"http": ip_addresses[i]}
        status = requests.get("http://httpbin.org/ip", proxies=proxy).status_code
        if status == 200:
            good.append(ip_addresses[i])
            print("GOOD IP : ", ip_addresses[i])
    except:
        if retry <= 2:
            retry = retry + 1
            continue
        dead.append(ip_addresses[i])
        print("DEAD IP : ", ip_addresses[i])
    i = i + 1

# saving good proxy to good proxy column
if len(good) > 0 :
    for row in range(len(good)):
        sheet.cell(row=row+2,column=2).value = good[row]

# saving good proxy to dead proxy column
if len(dead) > 0 :
    for row in range(len(dead)):
        sheet.cell(row=row+2,column=3).value = dead[row]

workbook.save("proxy_filter_file.xlsx")

